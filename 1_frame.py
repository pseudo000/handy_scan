from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from datetime import datetime
from PyQt5 import QtCore
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QTabWidget, QLabel, QPushButton, QLineEdit, QTreeWidget, QTreeWidgetItem, QHBoxLayout, QVBoxLayout, QStatusBar
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import (
    QApplication,
    QLabel,
    QLineEdit,
    QPushButton,
    QTabWidget,
    QTreeWidget,
    QVBoxLayout,
    QWidget,
    QHBoxLayout,
    QTreeWidgetItem,
)
from PyQt5 import QtCore, QtGui
from PyQt5.QtGui import QStandardItemModel, QStandardItem
import csv
from collections import defaultdict
import subprocess


class SpreadsheetGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Scan for SP IMPORT')
        self.previous_input = None
        self.pdf_path = None
        self.formatted_number = None

        # create tab widget
        self.tab_widget = QTabWidget()
        self.scan_tab = QWidget()
        self.inspection_tab = QWidget()
        self.sagawa_tab = QWidget()
        self.tab_widget.addTab(self.scan_tab, "EXCEL-CSV")

        # create widgets for scan tab
        self.filename_label = QLabel('No file selected')
        self.select_file_button = QPushButton('Select EXCEL File ■IMPORT■ (F2)')
        self.select_file_button.clicked.connect(self.select_file)

        self.select_file_button2 = QPushButton('Select EXCEL File ■EXPORT■ (F2)')
        self.select_file_button2.clicked.connect(self.select_file2)

        self.select_csv_button = QPushButton('Select CSV File')  # CSV 파일 선택 버튼 생성
        self.select_csv_button.clicked.connect(self.select_csv_file)  # 버튼 클릭 시 select_csv_file() 메서드 실행

        self.search_label = QLabel('검색어:')
        self.search_entry = QLineEdit()
        self.search_entry.returnPressed.connect(self.search_table)

        self.search_button = QPushButton('검색')
        self.search_button.clicked.connect(self.search_table)

        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(['NO.', 'HOUSE NO.', '荷主', '重量', '区分', 'CT', 'SCANNED', 'COMMENT'])
        self.tree.setColumnCount(8)

        # set column widths
        self.tree.setColumnWidth(0, 70)
        self.tree.setColumnWidth(1, 150)
        self.tree.setColumnWidth(2, 150)
        self.tree.setColumnWidth(3, 70)
        self.tree.setColumnWidth(4, 70)
        self.tree.setColumnWidth(5, 70)
        self.tree.setColumnWidth(6, 70)
        self.tree.setColumnWidth(7, 120)

        # set stylesheet to show grid lines
        self.tree.setStyleSheet(
            "QTreeView::item { border-bottom: 1px solid black; border-right: 1px solid black; }"
            "QTreeView::item:selected { background-color: #d8d8d8; }"
        )

        # enable sorting
        self.tree.setSortingEnabled(True)
        self.tree.sortItems(0, QtCore.Qt.AscendingOrder)

        # create a layout for search widgets
        search_layout = QHBoxLayout()
        search_layout.addWidget(self.search_label)
        search_layout.addWidget(self.search_entry)
        search_layout.addWidget(self.search_button)

        scan_layout = QHBoxLayout()

        left_widget = QWidget()
        left_layout = QVBoxLayout()

        button_layout = QHBoxLayout()  # 수평 레이아웃
        button_layout.addWidget(self.select_file_button)
        button_layout.addWidget(self.select_file_button2)

        left_layout.addWidget(self.filename_label)
        left_layout.addLayout(button_layout)
        left_layout.addWidget(self.select_csv_button)  # select_file_button 다음에 select_csv_button 추가
        left_layout.addSpacing(20)
        left_layout.addWidget(self.tree)
        left_layout.addSpacing(20)
        left_layout.addLayout(search_layout)

        self.search_results = []

        # create a save button and add it to the layout
        self.save_button = QPushButton('Save (F5)')
        self.save_button.clicked.connect(self.save_data)
        left_layout.addWidget(self.save_button)

        left_widget.setLayout(left_layout)

        scan_layout.addWidget(left_widget)
        self.scan_tab.setLayout(scan_layout)

        # create a status bar
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # add a custom widget to the status bar for displaying column sums
        self.column_sum_label = QLabel()
        self.status_bar.addPermanentWidget(self.column_sum_label)

        # connect signal to update column sums
        self.tree.itemSelectionChanged.connect(self.update_column_sums)

        # set initial column sums
        self.update_column_sums()

        # set layout
        layout = QVBoxLayout()
        layout.addWidget(self.tab_widget)

        main_widget = QWidget()
        main_widget.setLayout(layout)
        self.setCentralWidget(main_widget)
        self.setGeometry(100, 100, 1200, 800)

        font = QtGui.QFont('Meiryo')
        font.setPointSize(10)
        # font.setBold(True)

        # 전체적인 글자 크기를 조절할 위젯들에 폰트 설정
        self.tab_widget.setFont(font)
        self.filename_label.setFont(font)
        self.select_file_button.setFont(font)
        self.select_file_button2.setFont(font)
        self.select_csv_button.setFont(font)
        self.search_label.setFont(font)
        self.search_entry.setFont(font)
        self.search_button.setFont(font)
        self.tree.setFont(font)
        self.save_button.setFont(font)
        self.column_sum_label.setFont(font)

        background_color = QColor(230, 230, 230)  # 연한 하늘색
        self.setStyleSheet(f"QMainWindow {{ background-color: {background_color.name()}; }}")
        self.tree.setFrameStyle(QFrame.Box | QFrame.Plain)


    def search_table(self):
        search_text = self.search_entry.text().lower()
        self.search_results = []

        # traverse all rows in the table
        for row in range(self.tree.topLevelItemCount()):
            # traverse all columns in the row
            for column in range(self.tree.columnCount()):
                item = self.tree.topLevelItem(row).text(column).lower()
                if search_text in item:
                    self.search_results.append((row, column))

        if self.search_results:
            row, column = self.search_results[0]  # focus on the first result
            item = self.tree.topLevelItem(row)
            self.tree.setCurrentItem(item, column)

        self.search_entry.clear()

    def keyPressEvent(self, event):
        if event.key() == QtCore.Qt.Key_F5:
            self.save_data()
        if event.key() == QtCore.Qt.Key_F2:
            if self.select_file_button2.isChecked():
                self.select_file2()
            else:
                self.select_file()
        if event.key() == QtCore.Qt.Key_F4:
            self.back_one_step()

    def select_file(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx)")
        file_dialog.exec_()

        selected_files = file_dialog.selectedFiles()
        if selected_files:
            file_path = selected_files[0]
            self.filename_label.setText(file_path)
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            self.tree.clear()

            for i, row in enumerate(
                worksheet.iter_rows(min_row=2, min_col=1, max_col=22, values_only=True), start=1
            ):
                if row[2]:  # E열에 데이터가 있는 경우에만 해당 컬럼에 불러오기
                    item = QTreeWidgetItem(
                        [str(i), str(row[1]), str(row[5]), str(row[6]), str(row[19]), str(row[21])]
                    )
                    item.setText(1, str(row[5]))  # HOUSE NO.
                    item.setText(2, str(row[6]))  # 荷主
                    item.setText(3, str(row[21]))  # 重量
                    item.setText(4, str(row[1]))  # 区分
                    item.setText(5, str(row[19]))  # CT
                    self.tree.addTopLevelItem(item)

            self.update_column_sums()  # 상태바 업데이트 추가

    def select_file2(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("Excel Files (*.xlsx)")
        file_dialog.exec_()

        selected_files = file_dialog.selectedFiles()
        if selected_files:
            file_path = selected_files[0]
            self.filename_label.setText(file_path)
            workbook = load_workbook(file_path)
            worksheet = workbook.active

            self.tree.clear()

            for i, row in enumerate(
                worksheet.iter_rows(min_row=2, min_col=1, max_col=22, values_only=True), start=1
            ):
                if row[4]:  # E열에 데이터가 있는 경우에만 해당 컬럼에 불러오기
                    item = QTreeWidgetItem(
                        [str(i), str(row[1]), str(row[6]), str(row[9]), str(row[17]), str(row[18])]
                    )
                    item.setText(1, str(row[6]))  # HOUSE NO.
                    item.setText(2, str(row[9]))  # 荷主
                    item.setText(3, str(row[18]))  # 重量
                    item.setText(4, str(row[1]))  # 区分
                    item.setText(5, str(row[17]))  # CT
                    self.tree.addTopLevelItem(item)

            self.update_column_sums()  # 상태바 업데이트 추가


    def select_csv_file(self):
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilter("CSV 파일 (*.csv)")
        file_dialog.exec_()

        selected_files = file_dialog.selectedFiles()
        if selected_files:
            file_path = selected_files[0]
            self.filename_label.setText(file_path)

            with open(file_path, 'r') as file:
                reader = csv.reader(file)
                csv_data = []
                csv_sum_dict = defaultdict(int)

                for row in reader:
                    modified_row = [element.strip('a') if index == 4 else element for index, element in enumerate(row)]
                    csv_data.append(modified_row)

                    if modified_row[4] not in csv_sum_dict:
                        csv_sum_dict[modified_row[4]] = int(row[5])
                    else:
                        csv_sum_dict[modified_row[4]] += int(row[5])

            for i in range(self.tree.topLevelItemCount()):
                item = self.tree.topLevelItem(i)
                house_no = item.text(1)

                for row in csv_data:
                    modified_row = row[4]

                    if house_no == modified_row:
                        item.setText(6, row[5])

                        sum_value = csv_sum_dict[modified_row]
                        item.setText(6, str(sum_value))

                    if not item.text(6) or int(item.text(5)) > int(item.text(6)):
                        item.setText(7, "이상유무 확인 바랍니다.")
                    else:
                        item.setText(7, "OK")

                    item.setForeground(7, QtGui.QColor('red'))
                    font = QtGui.QFont()
                    font.setBold(True)
                    item.setFont(7, font)

            self.update_column_sums()
       

    def update_column_sums(self):
        # calculate column sums
        column_sum_3 = 0
        column_sum_5 = 0
        column_sum_6 = 0
        total_rows = self.tree.topLevelItemCount()  # 총 행 수 변수 추가
        scan_count = 0  # 스캔 건수 변수 추가

        for i in range(total_rows):
            item = self.tree.topLevelItem(i)
            column_sum_3 += float(item.text(3))
            column_sum_5 += float(item.text(5))
            if item.text(6):  # 7번째 컬럼 데이터 확인
                column_sum_6 += float(item.text(6))
                scan_count += 1

        # set column sum values to the status bar
        column_sum_text = f'<html><body><span style="font-weight: bold; color: red; font-size: x-large;">■총 건수: {total_rows}  ■총 CT수: {int(column_sum_5)}  ■총 중량: {column_sum_3:.2f}</span></body></html>'
        additional_text = f'<br/><span style="font-weight: bold; color: blue; font-size: x-large;">■스캔 건수: {scan_count}  ■스캔 CT수: {int(column_sum_6)}</span>'
        self.column_sum_label.setText(column_sum_text + additional_text)



    def save_data(self):
        # get file path and name
        file_path = self.filename_label.text()
        if not file_path:
            QMessageBox.warning(self, 'Error', 'Please select a file first.')
            return
        file_dir, file_name = os.path.split(file_path)
        file_name_no_ext, file_ext = os.path.splitext(file_name)

        # create new workbook and worksheet
        new_file_name = file_name_no_ext + '_new' + file_ext
        new_file_path, _ = QFileDialog.getSaveFileName(self, 'Save File As', new_file_name, 'Excel files (*.xlsx)')
        if not new_file_path:
            return
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Data'

        # write column headers
        headers = ['NO.', 'HOUSE NO.', '荷主', '重量', '区分', 'CT', 'SCANNED', 'COMMENT']
        for col_num, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_num)
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header

        # write data to worksheet
        for row_num in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(row_num)
            house_no_value = item.text(1)
            comment_value = item.text(7)

            row = [
                item.text(0),  # NO.
                house_no_value if house_no_value else '',  # HOUSE NO.
                '',  # 荷主 (빈 값)
                '',  # 重量 (빈 값)
                '',  # 区分 (빈 값)
                '',  # CT (빈 값)
                '',  # SCANNED (빈 값)
                comment_value if comment_value else ''  # COMMENT
            ]

            worksheet['A{}'.format(row_num + 2)] = row[1]  # 엑셀 A열에 저장
            worksheet['H{}'.format(row_num + 2)] = row[7]  # 엑셀 H열에 빈 값 저장
            worksheet['I{}'.format(row_num + 2)] = '0' if comment_value == 'OK' else '1'  # 엑셀 I열에 저장

        # save workbook to new file path
        workbook.save(new_file_path)
        QMessageBox.information(self, 'Saved', 'Data has been saved to {}.'.format(new_file_path))

        # open the saved file
        subprocess.Popen(['start', new_file_path], shell=True)



if __name__ == '__main__':
    import sys

    app = QApplication(sys.argv)
    spreadsheet_gui = SpreadsheetGUI()
    spreadsheet_gui.show()
    sys.exit(app.exec_())



# pyinstaller -w -F --icon="C:\Users\swwoo\Desktop\handy_csv\SCAN.ico" "C:\Users\swwoo\Desktop\handy_csv\1_frame.py"